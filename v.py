import pandas as pd
import streamlit as st

# Initialize session state for navigation and validations
if "validations" not in st.session_state:
    st.session_state.validations = []

if "current_index" not in st.session_state:
    st.session_state.current_index = 0

if "validation_started" not in st.session_state:
    st.session_state.validation_started = False


def main_page():
    """Main page for file upload and starting validation."""
    st.title("Valida las Cadenas asignadas a los restaurantes")

    # File upload
    uploaded_file = st.file_uploader(
        "Sube tu archivo (Excel o CSV)", type=["xlsx", "csv", "xls"])

    if uploaded_file:
        # Store the uploaded file name in session state
        st.session_state.df_name = uploaded_file.name

        # Determine the file type and load the DataFrame
        if uploaded_file.name.endswith(".csv"):
            df = pd.read_csv(uploaded_file)
        elif uploaded_file.name.endswith(".xlsx") or uploaded_file.name.endswith(".xls"):
            df = pd.read_excel(uploaded_file)
        else:
            st.error("Tipo de archivo no soportado.")
            return

        # Drop 'Unnamed: 0' column if it exists
        if 'Unnamed: 0' in df.columns:
            df = df.drop(columns=['Unnamed: 0'])

        # Show the first few rows of the DataFrame
        st.write("### Vista previa de los datos:")
        st.dataframe(df.head())

        # Add a button to start validation
        if st.button("Comenzar"):
            st.session_state.validation_started = True
            st.session_state.df = df  # Save DataFrame in session state


def validation_page():
    """Validation page for processing row-by-row."""
    st.title("Proceso de Validación")

    df = st.session_state.df  # Retrieve the stored DataFrame

    # Initialize validations if this is the first load
    if not st.session_state.validations or len(st.session_state.validations) != len(df):
        st.session_state.validations = ["No revisado"] * len(df)

    # Current row details
    current_index = st.session_state.current_index
    row = df.iloc[current_index]

    # Display row details
    st.write(f"### Validando fila {current_index + 1} de {len(df)}")
    st.write("---")

    # Use columns for a neat layout
    col1, col2 = st.columns([1, 2])

    # Key details on the left
    with col1:
        st.markdown(f"**Nombre Restaurante:**")
        st.info(row["cluster_name_clean"])

        st.markdown(f"**Cadena Asignada:**")
        st.info(row["normalized_main_chain"])

    # Additional details on the right
    with col2:
        st.markdown("**Otros Detalles:**")
        excluded_columns = ['cluster_name_clean', 'normalized_main_chain']
        for column, value in row.items():
            if column not in excluded_columns:
                st.markdown(f"- **{column}:** {value}")

    st.write("---")

    # Validation buttons
    # Validation buttons
    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("Correcto", key="boton_correcto"):
            st.session_state.validations[current_index] = "Correcto"
            if current_index < len(df) - 1:
                st.session_state.current_index += 1
                st.rerun()  # Updated from st.experimental_rerun()
    with col2:
        if st.button("Incorrecto", key="boton_incorrecto"):
            st.session_state.validations[current_index] = "Incorrecto"
            if current_index < len(df) - 1:
                st.session_state.current_index += 1
                st.rerun()  # Updated from st.experimental_rerun()
    with col3:
        if st.button("Atras", key="boton_atras") and current_index > 0:
            st.session_state.current_index -= 1
            st.rerun()  # Updated from st.experimental_rerun()

    # Save results button
    if st.button("Guardar Validación"):
        df["Validación"] = st.session_state.validations

        # Calculate percentages
        total_validations = len(df)
        correct_count = st.session_state.validations.count("Correcto")
        incorrect_count = st.session_state.validations.count("Incorrecto")

        correct_percentage = (
            correct_count / total_validations) * 100 if total_validations else 0
        incorrect_percentage = (
            incorrect_count / total_validations) * 100 if total_validations else 0

        if st.session_state.df_name.endswith(".csv"):
            # Save CSV as in-memory string
            validated_file = df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="Descargar Validación como CSV",
                data=validated_file,
                file_name=st.session_state.df_name.replace(
                    ".csv", "_validacion.csv"),
                mime="text/csv"
            )
        elif st.session_state.df_name.endswith((".xlsx", ".xls")):
            # Save Excel as in-memory bytes with formatting
            import io
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill

            # Create Excel writer with openpyxl engine
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Validaciones')

                # Access workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Validaciones']

                # Apply color formatting
                fill_green = PatternFill(
                    start_color="00FF00", end_color="00FF00", fill_type="solid")
                fill_red = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type="solid")

                # Start at row 2 (skip header)
                for row_idx, value in enumerate(df["Validación"], start=2):
                    cell = worksheet.cell(
                        row=row_idx, column=df.columns.get_loc("Validación") + 1)
                    if value == "Correcto":
                        cell.fill = fill_green
                    elif value == "Incorrecto":
                        cell.fill = fill_red

                # Add percentage calculations at the end of the sheet
                summary_row = len(df) + 3  # Leave one empty row after data
                worksheet.cell(row=summary_row, column=1,
                               value="Porcentaje Correctos:")
                worksheet.cell(row=summary_row, column=2,
                               value=f"{correct_percentage:.2f}%")
                worksheet.cell(row=summary_row + 1, column=1,
                               value="Porcentaje Incorrectos:")
                worksheet.cell(row=summary_row + 1, column=2,
                               value=f"{incorrect_percentage:.2f}%")

            # Save workbook to in-memory bytes
            validated_file = output.getvalue()
            st.download_button(
                label="Descargar Validación como Excel",
                data=validated_file,
                file_name=st.session_state.df_name.replace(
                    ".xlsx", "_validacion.xlsx"),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("Formato de archivo no soportado para guardar.")


if st.session_state.validation_started:
    validation_page()
else:
    main_page()


#        streamlit run val4.py
